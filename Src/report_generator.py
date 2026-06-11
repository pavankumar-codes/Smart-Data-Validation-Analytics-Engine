import openpyxl
import os

from exceptions import InvalidDatasetError

def generate_cleaned_valid_data_report(employee):

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Cleaned_Data"

    if len(employee) == 0:
        raise InvalidDatasetError("Cleaned Valid Dataset is Empty")

    sheet.append(list(employee[0].keys()))

    for emp in employee:
        sheet.append(list(emp.values()))

    folder_path = os.path.join(
        os.getcwd(),
        'Data',
        '2_Valid_Data'
    )

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    cleaned_valid_data_report_path = os.path.join(
        folder_path,
        'Cleaned_Valid_Data.xlsx'
    )

    wb.save(cleaned_valid_data_report_path)

def generate_invalid_data_report(employee):
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title = "Invalid_Data"

    if len(employee) == 0:
        raise InvalidDatasetError("InValid Dataset is Empty")
    
    header=list(employee[0]['employee'].keys())
    header.append('errors')
    sheet.append(header)

    for emp in employee:
        values=list(emp['employee'].values())
        error_string=", ".join(emp['errors'])
        values.append(error_string)
        sheet.append(values)
    #for emp in employee:
        #sheet.append(list(emp['employee'].values()))
    #sheet.cell(row=1,column=1,value="id")
    #sheet.cell(row=1,column=2,value="name")
    #sheet.cell(row=1,column=3,value="salary")
    #sheet.cell(row=1,column=4,value="department")
    #sheet.cell(row=1,column=5,value="experience")
    #sheet.cell(row=1,column=6,value="errors")
    #row=2
    #for emp in employee:
        #sheet.cell(row=row,column=1,value=emp['employee']['id'])
        #sheet.cell(row=row,column=2,value=emp['employee']['name'])
        #sheet.cell(row=row,column=3,value=emp['employee']['salary'])
        #sheet.cell(row=row,column=4,value=emp['employee']['department'])
        #sheet.cell(row=row,column=5,value=emp['employee']['experience'])
        #error_string=",".join(emp['errors'])
        #sheet.cell(row=row,column=6,value=error_string)
        #row+=1
    folder_path=os.path.join(os.getcwd(),'Data','3_Invalid_Data')
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    invalid_data_report_path=os.path.join(folder_path,'Invalid_Data.xlsx')
    
    wb.save(invalid_data_report_path)
   
def generate_global_kpi_report(global_kpi_report):
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title='Global_KPI_Report'
    
    if len(global_kpi_report)==0:
        raise InvalidDatasetError("Global KPI Dataset is Empty")
    sheet.cell(row=1,column=1,value='KPI')
    sheet.cell(row=1,column=2,value='Value')
    row=2
    row_column=1
    column=2
    for key,value in global_kpi_report.items():
        sheet.cell(row=row,column=row_column,value=key)
        sheet.cell(row=row,column=column,value=value)
        row+=1
    folder_path = os.path.join(
        os.getcwd(),
        'Data',
        '4_Data_Analytics'
    )

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    global_kpi_report_path = os.path.join(
        folder_path,
        'Global_KPI_Report.xlsx'
    )

    wb.save(global_kpi_report_path)

def generate_department_analytics_report(department_analytics_data):
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title='Department_Analytics_Report'
    if len(department_analytics_data) == 0:
        raise InvalidDatasetError("Department Analytics Dataset is Empty")
    sheet.cell(row=1,column=1,value='Department')
    column=2
    for data in department_analytics_data.values():

        for title in data.keys():
            sheet.cell(row=1,column=column,value=title)
            column+=1
        break
    
    row=2
    
    for key,value in department_analytics_data.items():
        column=1
        sheet.cell(row=row,column=column,value=key)
        column+=1
        for v in value.values():
            sheet.cell(row=row,column=column,value=v)
            column+=1
        row+=1
    
    folder_path=os.path.join(os.getcwd(),'Data','4_Data_Analytics')

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    department_analytics_data_report_path=os.path.join(folder_path,'Department_Analytics_Report.xlsx')


    wb.save(department_analytics_data_report_path)

def generate_business_insights_report(
        highest_paid_employee_data,
        lowest_paid_employee_data,
        most_experienced_employee_data,
        least_experienced_employee_data):

    business_insights = [
        ('Highest Paid Employee', highest_paid_employee_data),
        ('Lowest Paid Employee', lowest_paid_employee_data),
        ('Most Experienced Employee', most_experienced_employee_data),
        ('Least Experienced Employee', least_experienced_employee_data)
    ]

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Business_Insights_Report"

    if all(
        len(data) == 0
        for _, data in business_insights
    ):
        raise InvalidDatasetError(
            "Business Insights Dataset is Empty"
        )

    # Headers
    sheet.cell(row=1, column=1, value="Insight")

    title_list = list(highest_paid_employee_data[0].keys())

    column = 2
    for title in title_list:
        sheet.cell(
            row=1,
            column=column,
            value=title
        )
        column += 1

    # Data
    row = 2

    for insight, employee_data in business_insights:

        for employee in employee_data:

            sheet.cell(
                row=row,
                column=1,
                value=insight
            )

            column = 2

            for value in employee.values():

                sheet.cell(
                    row=row,
                    column=column,
                    value=value
                )

                column += 1

            row += 1

        # Leave one blank row between insight groups
        row += 1

    folder_path = os.path.join(
        os.getcwd(),
        'Data',
        '4_Data_Analytics'
    )

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    business_insights_report_path = os.path.join(
        folder_path,
        'Business_Insights_Report.xlsx'
    )

    wb.save(business_insights_report_path)

def generate_department_insights_report(
        department_with_highest_average_salary_data,
        department_with_lowest_average_salary_data,
        department_with_most_employees_data,
        department_with_least_employees_data):

    if (
        len(department_with_highest_average_salary_data) == 0 and
        len(department_with_lowest_average_salary_data) == 0 and
        len(department_with_most_employees_data) == 0 and
        len(department_with_least_employees_data) == 0
    ):
        raise InvalidDatasetError(
            "Department Insights Dataset is Empty"
        )

    department_insights = [
        (
            'Department With Highest Average Salary',
            department_with_highest_average_salary_data
        ),
        (
            'Department With Lowest Average Salary',
            department_with_lowest_average_salary_data
        ),
        (
            'Department With Most Employees',
            department_with_most_employees_data
        ),
        (
            'Department With Least Employees',
            department_with_least_employees_data
        )
    ]

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Department_Insights_Report'

    sheet.append([
        'Insight',
        'Department'
    ])

    for insight, departments in department_insights:

        department_string = ", ".join(departments)

        sheet.append([
            insight,
            department_string
        ])

    folder_path = os.path.join(
        os.getcwd(),
        'Data',
        '4_Data_Analytics'
    )

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    department_insights_report_path = os.path.join(
        folder_path,
        'Department_Insights_Report.xlsx'
    )

    wb.save(department_insights_report_path)

def generate_salary_distribution_analytics_report(
        salary_distribution_analytics_data):

    wb = openpyxl.Workbook()

    # ==================================================
    # Summary Sheet
    # ==================================================

    summary_sheet = wb.active
    summary_sheet.title = "Salary_Summary"

    low_salary_employees = salary_distribution_analytics_data[
        'low_salary_employees'
    ]

    mid_salary_employees = salary_distribution_analytics_data[
        'mid_salary_employees'
    ]

    high_salary_employees = salary_distribution_analytics_data[
        'high_salary_employees'
    ]

    summary_sheet.append([
        'Insight',
        'Count'
    ])

    salary_summary = [
        (
            'Low Salary Employee Count',
            len(low_salary_employees)
        ),
        (
            'Mid Salary Employee Count',
            len(mid_salary_employees)
        ),
        (
            'High Salary Employee Count',
            len(high_salary_employees)
        )
    ]

    for insight, count in salary_summary:

        summary_sheet.append([
            insight,
            count
        ])

    # ==================================================
    # Low Salary Employees Sheet
    # ==================================================

    low_salary_sheet = wb.create_sheet(
        'Low_Salary_Employees'
    )

    if len(low_salary_employees) > 0:

        low_salary_sheet.append(
            list(low_salary_employees[0].keys())
        )

        for employee in low_salary_employees:

            low_salary_sheet.append(
                list(employee.values())
            )

    # ==================================================
    # Mid Salary Employees Sheet
    # ==================================================

    mid_salary_sheet = wb.create_sheet(
        'Mid_Salary_Employees'
    )

    if len(mid_salary_employees) > 0:

        mid_salary_sheet.append(
            list(mid_salary_employees[0].keys())
        )

        for employee in mid_salary_employees:

            mid_salary_sheet.append(
                list(employee.values())
            )

    # ==================================================
    # High Salary Employees Sheet
    # ==================================================

    high_salary_sheet = wb.create_sheet(
        'High_Salary_Employees'
    )

    if len(high_salary_employees) > 0:

        high_salary_sheet.append(
            list(high_salary_employees[0].keys())
        )

        for employee in high_salary_employees:

            high_salary_sheet.append(
                list(employee.values())
            )

    # ==================================================
    # Save Report
    # ==================================================

    folder_path = os.path.join(
        os.getcwd(),
        'Data',
        '4_Data_Analytics'
    )

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    salary_distribution_report_path = os.path.join(
        folder_path,
        'Salary_Distribution_Report.xlsx'
    )

    wb.save(
        salary_distribution_report_path
    )

def generate_executive_summary_report(
        valid_data,
        invalid_data,
        kpi_data,
        department_analytics_data):

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Executive_Summary_Report"

    total_records = len(valid_data) + len(invalid_data)
    valid_records = len(valid_data)
    invalid_records = len(invalid_data)

    validation_success_rate = round(
        (valid_records / total_records) * 100,
        2
    ) if total_records > 0 else 0

    executive_summary = {
        "Total Records": total_records,
        "Valid Records": valid_records,
        "Invalid Records": invalid_records,
        "Validation Success Rate (%)": validation_success_rate,
        "Highest Salary": kpi_data["highest_salary"],
        "Lowest Salary": kpi_data["lowest_salary"],
        "Average Salary": kpi_data["average_salary"],
        "Highest Experience": kpi_data["highest_experience"],
        "Lowest Experience": kpi_data["lowest_experience"],
        "Average Experience": kpi_data["average_experience"],
        "Department Count": len(department_analytics_data)
    }

    sheet.cell(row=1, column=1, value="Metric")
    sheet.cell(row=1, column=2, value="Value")

    row = 2

    for key, value in executive_summary.items():

        sheet.cell(
            row=row,
            column=1,
            value=key
        )

        sheet.cell(
            row=row,
            column=2,
            value=value
        )

        row += 1

    folder_path = os.path.join(
        os.getcwd(),
        'Data',
        '4_Data_Analytics'
    )

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    executive_summary_report_path = os.path.join(
        folder_path,
        'Executive_Summary_Report.xlsx'
    )

    wb.save(executive_summary_report_path)
